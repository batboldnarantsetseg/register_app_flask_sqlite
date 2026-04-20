import os
import re
import sqlite3
from io import BytesIO
from functools import wraps
from datetime import datetime

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    send_file,
    g,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_DIR = os.path.join(BASE_DIR, "instance")
os.makedirs(DB_DIR, exist_ok=True)
DATABASE = os.path.join(DB_DIR, "register.db")

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "change-this-secret-key")
app.config["ADMIN_USERNAME"] = os.environ.get("ADMIN_USERNAME", "admin")
app.config["ADMIN_PASSWORD"] = os.environ.get("ADMIN_PASSWORD", "admin123")

EMAIL_REGEX = re.compile(r"^[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}$")
PHONE_REGEX = re.compile(r"^[0-9+\-\s]{8,20}$")


def get_db():
    if "db" not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
    return g.db


def init_db():
    db = sqlite3.connect(DATABASE)
    cursor = db.cursor()
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS Register (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            last_name TEXT NOT NULL,
            first_name TEXT NOT NULL,
            phone TEXT NOT NULL,
            email TEXT NOT NULL UNIQUE,
            created_at TEXT NOT NULL
        )
        """
    )
    db.commit()
    db.close()


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def admin_required(view):
    @wraps(view)
    def wrapped_view(*args, **kwargs):
        if not session.get("admin_logged_in"):
            flash("Админ хуудсанд нэвтрэх шаардлагатай.", "warning")
            return redirect(url_for("admin_login"))
        return view(*args, **kwargs)

    return wrapped_view


@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        last_name = request.form.get("last_name", "").strip()
        first_name = request.form.get("first_name", "").strip()
        phone = request.form.get("phone", "").strip()
        email = request.form.get("email", "").strip().lower()

        errors = []
        if not last_name:
            errors.append("Овог оруулна уу.")
        if not first_name:
            errors.append("Нэр оруулна уу.")
        if not phone:
            errors.append("Утасны дугаар оруулна уу.")
        elif not PHONE_REGEX.match(phone):
            errors.append("Утасны дугаарын формат буруу байна.")
        if not email:
            errors.append("Имэйл оруулна уу.")
        elif not EMAIL_REGEX.match(email):
            errors.append("Имэйл хаяг зөв форматтай биш байна. Жишээ: example@company.com")

        if errors:
            for error in errors:
                flash(error, "danger")
            return render_template("index.html")

        try:
            db = get_db()
            db.execute(
                "INSERT INTO Register (last_name, first_name, phone, email, created_at) VALUES (?, ?, ?, ?, ?)",
                (last_name, first_name, phone, email, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            )
            db.commit()
            flash("Бүртгэл амжилттай хийгдлээ.", "success")
            return redirect(url_for("index"))
        except sqlite3.IntegrityError:
            flash("Энэ имэйл хаяг өмнө нь бүртгэгдсэн байна.", "danger")
            return render_template("index.html")

    return render_template("index.html")


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if (
            username == app.config["ADMIN_USERNAME"]
            and password == app.config["ADMIN_PASSWORD"]
        ):
            session["admin_logged_in"] = True
            flash("Админ хэсэгт амжилттай нэвтэрлээ.", "success")
            return redirect(url_for("admin_dashboard"))
        else:
            flash("Нэвтрэх нэр эсвэл нууц үг буруу байна.", "danger")

    return render_template("admin_login.html")


@app.route("/admin/logout")
def admin_logout():
    session.clear()
    flash("Системээс гарлаа.", "info")
    return redirect(url_for("admin_login"))


@app.route("/admin")
@admin_required
def admin_dashboard():
    db = get_db()
    registrations = db.execute(
        "SELECT id, last_name, first_name, phone, email, created_at FROM Register ORDER BY id DESC"
    ).fetchall()
    return render_template("admin_dashboard.html", registrations=registrations)


@app.route("/admin/export")
@admin_required
def export_excel():
    db = get_db()
    rows = db.execute(
        "SELECT id, last_name, first_name, phone, email, created_at FROM Register ORDER BY id DESC"
    ).fetchall()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Register"

    headers = ["ID", "Овог", "Нэр", "Утас", "Имэйл", "Бүртгэгдсэн огноо"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="203A43")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        sheet.append(
            [
                row["id"],
                row["last_name"],
                row["first_name"],
                row["phone"],
                row["email"],
                row["created_at"],
            ]
        )

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    column_widths = {
        "A": 8,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 32,
        "F": 22,
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"register_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    init_db()
    app.run(host='0.0.0.0', debug=True)
else:
    init_db()
